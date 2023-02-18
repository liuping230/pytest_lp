#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/2/17 13:39
# @Author  : being℡
# @File    : test_excel_v1.py
# @Software: python3.9.5 win10
import openpyxl
import pytest

from pytest_excel_allure.api_keyword.api_key import ApiKey
from pytest_excel_allure.data_driver.read_excel import read_excel


# setup_module:在当前文件中，在所有用例执行之前执行
def setup_module():
    # 1、定义全局变量
    global ak, excel, sheet, excel_path
    # 2、实例化工具类
    ak = ApiKey()
    # 3、初始化Excel文件
    excel_path = './data/testdata.xlsx'
    excel =openpyxl.load_workbook(excel_path)
    sheet = excel['Sheet1']

@pytest.mark.parametrize('data',read_excel())
def test01(data):
    # 如果存在请求头
    assert_value = data[7]
    expect_value = data[8]
    request_value = data[5]
    r = data[0]
    if data[4]:
        # 存在请求参数
        if request_value:
            dict_data = {
                'url': data[1] + data[2],
                # eval函数官方解释：将字符串当做有效的表达式来求值并返回结算结果
                # 这里直接给headers 一个字典值
                'headers': eval(data[4]),
                # value[6]参数类型，data请求参数
                data[6]: eval(request_value)
            }
        # 不存在请求参数
        else:
            dict_data = {
                'url': data[1] + data[2],
                # eval函数官方解释：将字符串当做有效的表达式来求值并返回结算结果
                # 这里直接给headers 一个字典值
                'headers': eval(data[4]),
            }
    # 不存在请求头
    else:
        # 存在请求参数
        if request_value:
            dict_data = {
                'url': data[1] + data[2],
                # eval函数官方解释：将字符串当做有效的表达式来求值并返回结算结果
                # 这里直接给headers 一个字典值
                # value[6]参数类型，data请求参数
                data[6]: eval(request_value)
            }
        # 不存在请求参数
        else:
            dict_data = {
                'url': data[1] + data[2]
            }
        # print(dict_data)

        # 常规的参数传递：
        # ak.post(url=dict_data['url'],json = dict_data[value[6]])
        # 需要判断请求方法和请求参数信息...

        # 用反射来进行模拟请求
        res = getattr(ak, data[3])(**dict_data)
        # print('\n'+'----------响应结果---------')
        # print(res.text)

        # 结果校验
        try:
            result = ak.get_text(res.text, assert_value)
            # print('++++提取结果+++++')
            # print(result)
            if result == expect_value:
                sheet.cell(r, 10).value = '通过'
            else:
                sheet.cell(r, 10).value = '不通过'
            excel.save(excel_path)

        except:
            # print('++++++++++++实际结果++++++++++++++++++++')
            # print('请求参数有问题，请检查Excel表格数据')
            sheet.cell(r, 10).value = '请求参数有问题，请检查Excel表格数据'
            excel.save(excel_path)





if __name__ == '__main__':
    pytest.main(['-vs','test_excel_v2.py'])
